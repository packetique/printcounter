# version 1.0.5
import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from bs4 import BeautifulSoup
import requests
import shutil
import urllib3
from selenium import webdriver
URL400 = "/info_configuration.html?tab=Home&menu=DevConfig"
URL600 = "/hp/device/InternalPages/Index?id=UsagePage"
date = str(datetime.date.today().strftime("%d.%m.%Y"))


def main():

    def get_link(ip, model):
        if "425" in model or "426" in model:
            return 'http://' + str(ip) + URL400
        elif "602" in model or "605" in model or "750" in model:
            return 'https://' + str(ip) + URL600
        else:
            return 'model error'

    def do_400():
        td = soup.find("td", text="Всего оттисков:")
        tr = td.find_parent('tr')
        counter = int(tr.findChildren("td")[1].get_text())
        return counter

    def do_600():
        td = soup.find("td", id="UsagePage.EquivalentImpressionsTable.Print.Total")
        total = td.get_text()
        total = total[0:-1]
        counter = int(total.replace(',', '').replace('.', ''))
        return counter

    def check_output_dir(output_dir):
        if os.path.exists(output_dir):
            print('Директория {} существует, будет удалена и пересоздана'.format(output_dir))
            shutil.rmtree(output_dir)
        try:
            print('Cоздание директории {}...'.format(output_dir))
            os.mkdir(output_dir)
            print('Директория {} успешно создана'.format(output_dir))
        except Exception as ex:
            print('Ошибка создания директории')
            print(str(ex.__class__))

    check_output_dir(date)
    destination_filename = date + "\\ПКС РЦ Уткина Заводь " + date + ".xlsx"
    source_filename = "default.xlsx"
    wb = load_workbook(source_filename)
    sheet = wb.active
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    options = webdriver.ChromeOptions()
    options.add_argument("--ignore-certificate-errors-spki-list")
    driver = webdriver.Chrome("chromedriver.exe", chrome_options=options)
    driver.maximize_window()
    print('Всего записей: ' + str(sheet.max_row - 1))
    print()
    for row in range(2, sheet.max_row+1, 1):
        ip = sheet.cell(row, 1).value
        model = sheet.cell(row, 2).value
        myurl = get_link(ip, model)
        try:
            page = requests.get(myurl, verify=False, timeout=10)
            soup = BeautifulSoup(page.text, 'html.parser')
            if URL400 in myurl:
                counter = do_400()
            elif URL600 in myurl:
                counter = do_600()
            else:
                counter = 'error'
            savescreen = date + "\\" + ip + ".png"
            driver.get(myurl)
            element = driver.find_element_by_tag_name('body')
            element.screenshot(savescreen)
            sheet.cell(row, 5, counter)
            print(ip, counter)
        except Exception as e:
            counter = str(e.__class__)
            counter = counter.replace("<class 'requests.exceptions.", '').replace("'>", '')
            myfill = PatternFill(fgColor=RED, fill_type='solid')
            sheet.cell(row, 5).fill = myfill
            sheet.cell(row, 5, counter)
            print(ip, "Fail", "Вручную")

    driver.quit()
    sheet.cell(1, 5, "счетчик ("+date+")")
    wb.save(destination_filename)
    print()
    print('Выполнение программы завершено')


if __name__ == "__main__":
    main()
